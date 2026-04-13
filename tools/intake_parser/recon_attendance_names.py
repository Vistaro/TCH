"""
Build a name-reconciliation workbook between the Tuniti attendance xlsx
(one sheet per cohort 1-9) and the LIVE canonical student names from
the production database.

Pre-req: live students dumped to /tmp/students_live.tsv via:
  ssh -i ".../intelligentae_deploy_ed25519" intelligentae.co.uk@ssh.gb.stackcp.com \
    "mysql -h ... -B -e \"SELECT p.tch_id, p.full_name, s.cohort
       FROM persons p JOIN students s ON s.person_id=p.id
       ORDER BY s.cohort, p.full_name;\""

Output: one xlsx, one tab per cohort. Columns:
  A = name as it appears in the attendance spreadsheet
  B = best guess at the canonical name in the system (with TCH ID)
  C = blank — Ross fills if Col B is wrong
  D = match confidence
"""
from __future__ import annotations
from difflib import SequenceMatcher
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ATTENDANCE_XLSX = Path(r"C:/Users/Intel/AppData/Local/Temp/Ross Intake 1-9 (3).xlsx")
LIVE_DUMP = Path(r"C:/Users/Intel/AppData/Local/Temp/students_live.tsv")
OUT_PATH = Path(r"C:/ClaudeCode/_global/output/TCH/Tuniti Attendance Name Recon Apr-26.xlsx")


def load_live_students() -> dict[int, list[tuple[str, str]]]:
    """Returns {cohort_num: [(canonical_name, tch_id), ...]}."""
    out: dict[int, list[tuple[str, str]]] = {}
    for line in LIVE_DUMP.read_text(encoding="utf-8").splitlines()[1:]:
        parts = line.split("\t")
        if len(parts) < 3:
            continue
        tch_id, full_name, cohort = parts[0].strip(), parts[1].strip(), parts[2].strip()
        if not cohort.lower().startswith("cohort "):
            continue
        try:
            n = int(cohort.split()[1])
        except (IndexError, ValueError):
            continue
        out.setdefault(n, []).append((full_name, tch_id))
    return out


def attendance_names_per_sheet() -> dict[int, list[str]]:
    wb = load_workbook(ATTENDANCE_XLSX, data_only=True)
    out = {}
    for i, ws in enumerate(wb.worksheets, start=1):
        names = []
        for row in range(3, ws.max_row + 1):
            v = ws.cell(row=row, column=1).value
            if v and isinstance(v, str) and v.strip():
                names.append(v.strip())
        out[i] = names
    return out


def best_match(name: str, candidates: list[tuple[str, str]]) -> tuple[str, str, float]:
    if not candidates:
        return ("", "", 0.0)
    n = name.lower().strip()
    scored = [(c, t, SequenceMatcher(None, n, c.lower().strip()).ratio()) for c, t in candidates]
    scored.sort(key=lambda x: x[2], reverse=True)
    return scored[0]


def main():
    attendance = attendance_names_per_sheet()
    live = load_live_students()
    out = Workbook()
    out.remove(out.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    weak_fill = PatternFill("solid", fgColor="FFF2CC")
    none_fill = PatternFill("solid", fgColor="F8CBAD")

    for cohort in sorted(attendance.keys()):
        canon = live.get(cohort, [])
        ws = out.create_sheet(f"Cohort {cohort}")
        ws.append([
            "Name in attendance sheet",
            "Best guess at canonical name (Claude)",
            "Confirmed canonical name (Ross — fill only if B wrong)",
            "Confidence",
        ])
        for col_idx in range(1, 5):
            c = ws.cell(row=1, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="left", vertical="center")

        for src_name in attendance[cohort]:
            match, tch_id, score = best_match(src_name, canon)
            pct = round(score * 100)
            display = f"{match}  [{tch_id}]" if match else ""
            ws.append([src_name, display, "", f"{pct}%"])
            r = ws.max_row
            if score == 0:
                ws.cell(row=r, column=2).fill = none_fill
            elif score < 0.85:
                ws.cell(row=r, column=2).fill = weak_fill

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 48
        ws.column_dimensions["C"].width = 38
        ws.column_dimensions["D"].width = 12
        ws.freeze_panes = "A2"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    out.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    for cohort in sorted(attendance.keys()):
        print(f"  Cohort {cohort}: {len(attendance[cohort])} attendance names "
              f"vs {len(live.get(cohort, []))} live students")


if __name__ == "__main__":
    main()
