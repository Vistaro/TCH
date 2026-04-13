"""
Import the Tuniti attendance + course-summary spreadsheet
("Ross Intake 1-9 (3).xlsx") into the live TCH system.

For each student in cohorts 1-9 we write:

  Student record fields (table `students`):
    avg_score          ← Average column (already a 0..1 decimal)
    practical_status   ← Practical Placement column (facility name or "N/A")
    qualified          ← Qualified column (final status text)

  Person record fields (table `persons`):
    none — all training data lives on `students`

  Enrollment notes:
    `students.import_notes` is left alone — historical migrated already.
    Each value lands as a Note entry in the timeline tagged
      source       = 'import'
      source_ref   = 'Ross Intake 1-9 (3).xlsx#Cohort N!CELL'
      source_batch = 'tuniti-attendance-2026-04-13'

  Per-week attendance (table `training_attendance`):
    one row per (student × week) where the cell colour is green or red
      attendance_date = the Monday from row 1
      attendance_type = derived from module name (induction/module=>classroom,
                        practical=>practical, exam=>classroom)
      hours = NULL (cell only carries P/A)
      notes = "<Module name> — Present|Absent (Cohort N!CELL)"

Outputs:
  - SQL file written to /tmp/import_attendance.sql for Ross to apply
  - Per-cohort summary printed to stdout

Name mapping uses the LIVE persons.cohort + students.cohort match.
The two manual overrides Ross confirmed are baked in:
  Cohort 1 "Nelly"               → TCH-000003 Nelly Nachilongo
  Cohort 1 "Wisani Precious Mash" → TCH-000008 Wisani Precious Mashaba
"""
from __future__ import annotations
import csv
import re
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ATTENDANCE_XLSX = Path(r"C:/Users/Intel/AppData/Local/Temp/Ross Intake 1-9 (3).xlsx")
LIVE_DUMP       = Path(r"C:/Users/Intel/AppData/Local/Temp/students_live.tsv")
OUT_SQL         = Path(r"C:/Users/Intel/AppData/Local/Temp/import_attendance.sql")
SOURCE_FILE     = "Ross Intake 1-9 (3).xlsx"
BATCH_TAG       = "tuniti-attendance-2026-04-13"

# ── Manual name overrides (Ross-confirmed) ────────────────────────────
MANUAL_OVERRIDES = {
    (1, "Nelly"):                   "TCH-000003",
    (1, "Wisani Precious Mash"):    "TCH-000008",
}

# ── Column layout per sheet (1-based) ─────────────────────────────────
# Sheets 1-8: weekly cols B..K (2..11), summary M..S (13..19)
# Sheet 9 has one extra weekly column (B..L = 2..12), summary N..T (14..20)
SUMMARY_COLS_18 = {"sponsored": 13, "average": 14, "placement": 15, "type": 16,
                   "attendance": 17, "hours_left": 18, "qualified": 19}
SUMMARY_COLS_9  = {"sponsored": 14, "average": 15, "placement": 16, "type": 17,
                   "attendance": 18, "hours_left": 19, "qualified": 20}
WEEK_RANGE_18 = (2, 11)   # cols B..K
WEEK_RANGE_9  = (2, 12)   # cols B..L

# ── Colour → P/A mapping ──────────────────────────────────────────────
GREEN_RGB = {"FF92D050", "FF00B050"}      # bright greens
RED_RGB   = {"FFFF0000", "FFC00000", "FFFF1F1F"}
GREEN_THEMES = {9}  # theme index 9 with any positive tint = a green
PRESENT = "P"
ABSENT  = "A"


def cell_pa(cell) -> str | None:
    fg = cell.fill.fgColor
    if fg.type == "rgb":
        rgb = (fg.value or "").upper()
        if rgb in GREEN_RGB:
            return PRESENT
        if rgb in RED_RGB:
            return ABSENT
        return None
    if fg.type == "theme":
        if fg.value in GREEN_THEMES and (fg.tint or 0) > 0:
            return PRESENT
    return None


def attendance_type_for_module(name: str) -> str:
    n = (name or "").lower()
    if "practical" in n and "exam" not in n:
        return "practical"
    if "ojt" in n or "on the job" in n:
        return "ojt"
    return "classroom"


def excel_serial_to_date(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, (int, float)):
        # Excel epoch 1899-12-30 (with Lotus bug)
        d = datetime(1899, 12, 30) + timedelta(days=int(v))
        return d.date().isoformat()
    if isinstance(v, str):
        return v.strip() or None
    return None


def load_live_students() -> dict[int, list[tuple[str, str, int]]]:
    """{cohort_num: [(canonical_name, tch_id, person_id), ...]}."""
    out: dict[int, list[tuple[str, str, int]]] = {}
    for line in LIVE_DUMP.read_text(encoding="utf-8").splitlines()[1:]:
        parts = line.split("\t")
        if len(parts) < 3:
            continue
        tch_id, full_name, cohort = parts[0].strip(), parts[1].strip(), parts[2].strip()
        if not cohort.lower().startswith("cohort "):
            continue
        try:
            n = int(cohort.split()[1])
        except (IndexError, ValueError):
            continue
        try:
            pid = int(tch_id.replace("TCH-", "").lstrip("0") or "0")
        except ValueError:
            continue
        out.setdefault(n, []).append((full_name, tch_id, pid))
    return out


def map_xlsx_name_to_person(cohort: int, xlsx_name: str,
                            live: list[tuple[str, str, int]]) -> tuple[int, str, str] | None:
    """Returns (person_id, tch_id, canonical_name) or None."""
    if (cohort, xlsx_name) in MANUAL_OVERRIDES:
        target = MANUAL_OVERRIDES[(cohort, xlsx_name)]
        for fn, tch, pid in live:
            if tch == target:
                return (pid, tch, fn)
        return None
    n = xlsx_name.lower().strip()
    scored = [(fn, tch, pid, SequenceMatcher(None, n, fn.lower()).ratio()) for fn, tch, pid in live]
    scored.sort(key=lambda x: x[3], reverse=True)
    if scored and scored[0][3] >= 0.85:
        return (scored[0][2], scored[0][1], scored[0][0])
    return None


def sql_str(s):
    if s is None or s == "":
        return "NULL"
    if isinstance(s, (int, float)):
        return str(s)
    return "'" + str(s).replace("\\", "\\\\").replace("'", "''") + "'"


def main():
    wb = load_workbook(ATTENDANCE_XLSX, data_only=True)
    live = load_live_students()

    sql = []
    sql.append("-- AUTO-GENERATED Tuniti attendance + summary import")
    sql.append(f"-- Source: {SOURCE_FILE}")
    sql.append(f"-- Batch:  {BATCH_TAG}")
    sql.append(f"-- Generated: {datetime.now().isoformat()}")
    sql.append("START TRANSACTION;")
    sql.append("")

    summary_log = []

    for sheet_idx, ws in enumerate(wb.worksheets, start=1):
        is_sheet9 = (sheet_idx == 9)
        cols = SUMMARY_COLS_9 if is_sheet9 else SUMMARY_COLS_18
        wk_start, wk_end = (WEEK_RANGE_9 if is_sheet9 else WEEK_RANGE_18)

        # Pull week dates from row 1, module names from row 2
        weeks = []
        for c in range(wk_start, wk_end + 1):
            d = excel_serial_to_date(ws.cell(row=1, column=c).value)
            m = ws.cell(row=2, column=c).value
            weeks.append((c, d, str(m).strip() if m else f"Week col {get_column_letter(c)}"))

        live_cohort = live.get(sheet_idx, [])
        ws_count = 0
        ws_atten = 0
        ws_summary = 0

        # Find latest enrollment id per person (one query batch needed; we'll resolve at SQL apply
        # time using a subselect — keeps the script DB-free)

        for row in range(3, ws.max_row + 1):
            xlsx_name = ws.cell(row=row, column=1).value
            if not (xlsx_name and isinstance(xlsx_name, str) and xlsx_name.strip()):
                continue
            xlsx_name = xlsx_name.strip()
            mapped = map_xlsx_name_to_person(sheet_idx, xlsx_name, live_cohort)
            if not mapped:
                summary_log.append(f"Cohort {sheet_idx} row {row} '{xlsx_name}': NO MATCH — skipped")
                continue
            person_id, tch_id, canon = mapped
            ws_count += 1

            # ── Summary fields → students UPDATE + Notes ──
            updates = {}
            notes_to_add = []  # list of (subject, body, source_ref)

            avg_v = ws.cell(row=row, column=cols["average"]).value
            if isinstance(avg_v, (int, float)) and avg_v > 0:
                updates["avg_score"] = round(float(avg_v), 4)
                cell_addr = f"{get_column_letter(cols['average'])}{row}"
                notes_to_add.append((
                    "Average score set",
                    f"Set to {round(float(avg_v) * 100, 1)}% from Tuniti spreadsheet.",
                    f"{SOURCE_FILE}#{ws.title.strip()}!{cell_addr}"
                ))

            placement_v = ws.cell(row=row, column=cols["placement"]).value
            if placement_v and str(placement_v).strip() and str(placement_v).strip().lower() != 'n/a':
                updates["practical_status"] = str(placement_v).strip()[:100]
                cell_addr = f"{get_column_letter(cols['placement'])}{row}"
                notes_to_add.append((
                    "Practical placement set",
                    f"OJT placement: {placement_v}",
                    f"{SOURCE_FILE}#{ws.title.strip()}!{cell_addr}"
                ))

            qual_v = ws.cell(row=row, column=cols["qualified"]).value
            if qual_v and str(qual_v).strip():
                updates["qualified"] = str(qual_v).strip()[:50]
                cell_addr = f"{get_column_letter(cols['qualified'])}{row}"
                notes_to_add.append((
                    "Qualified status set",
                    f"Qualified status: {qual_v}",
                    f"{SOURCE_FILE}#{ws.title.strip()}!{cell_addr}"
                ))

            # Sponsored / Type / Attendance / Hours Left → only Notes (no column on students)
            for label, key, subject_tpl in [
                ("Sponsored amount",   "sponsored",  "Sponsored amount"),
                ("Type / care kind",   "type",       "Care type"),
                ("OJT attendance",     "attendance", "OJT attendance status"),
                ("OJT hours left",     "hours_left", "OJT hours left"),
            ]:
                v = ws.cell(row=row, column=cols[key]).value
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                cell_addr = f"{get_column_letter(cols[key])}{row}"
                notes_to_add.append((
                    subject_tpl,
                    f"{label}: {v}",
                    f"{SOURCE_FILE}#{ws.title.strip()}!{cell_addr}"
                ))

            if updates:
                set_clause = ", ".join(f"{k} = {sql_str(v)}" for k, v in updates.items())
                sql.append(f"UPDATE students SET {set_clause} WHERE person_id = {person_id};")
                ws_summary += 1

            for subject, body, src_ref in notes_to_add:
                sql.append(
                    "INSERT INTO activities "
                    "(activity_type_id, entity_type, entity_id, user_id, subject, notes, "
                    "source, source_ref, source_batch, activity_date, is_task, task_status) VALUES "
                    f"(7, 'persons', {person_id}, 3, {sql_str(subject)}, {sql_str(body)}, "
                    f"'import', {sql_str(src_ref)}, {sql_str(BATCH_TAG)}, NOW(), 0, 'pending');"
                )

            # ── Weekly attendance grid ──
            for col, week_date, module in weeks:
                cell = ws.cell(row=row, column=col)
                pa = cell_pa(cell)
                if pa is None:
                    continue
                cell_addr = f"{get_column_letter(col)}{row}"
                src_ref = f"{SOURCE_FILE}#{ws.title.strip()}!{cell_addr}"
                atten_type = attendance_type_for_module(module)
                # Map P/A onto a meaningful row in training_attendance.
                # We DON'T want to dedupe here — the apply-side ON DUPLICATE KEY would need a unique
                # constraint we don't have. The script is run once.
                pa_label = "Present" if pa == PRESENT else "Absent"
                notes_text = f"{module} — {pa_label} (from {ws.title.strip()}!{cell_addr})"
                # Look up the student's enrollment id at apply time via subselect
                date_sql = sql_str(week_date) if week_date else "NULL"
                sql.append(
                    "INSERT INTO training_attendance "
                    "(student_person_id, enrollment_id, attendance_date, attendance_type, hours, notes) "
                    f"SELECT {person_id}, "
                    f"(SELECT id FROM student_enrollments WHERE student_person_id = {person_id} "
                    f"ORDER BY enrolled_at DESC LIMIT 1), "
                    f"{date_sql}, '{atten_type}', NULL, {sql_str(notes_text)};"
                )
                # Mirror Note in the timeline
                sql.append(
                    "INSERT INTO activities "
                    "(activity_type_id, entity_type, entity_id, user_id, subject, notes, "
                    "source, source_ref, source_batch, activity_date, is_task, task_status) VALUES "
                    f"(7, 'persons', {person_id}, 3, "
                    f"{sql_str(f'Attendance: {module} ({pa_label})')}, "
                    f"{sql_str(notes_text)}, "
                    f"'import', {sql_str(src_ref)}, {sql_str(BATCH_TAG)}, "
                    f"{date_sql if week_date else 'NOW()'}, 0, 'pending');"
                )
                ws_atten += 1

        summary_log.append(f"Cohort {sheet_idx}: {ws_count} students, "
                          f"{ws_summary} student rows updated, {ws_atten} attendance rows")

    sql.append("COMMIT;")
    OUT_SQL.write_text("\n".join(sql), encoding="utf-8")

    print("\n".join(summary_log))
    print(f"\nSQL file: {OUT_SQL}")
    print(f"Total SQL statements: {sum(1 for l in sql if l.endswith(';'))}")


if __name__ == "__main__":
    main()
